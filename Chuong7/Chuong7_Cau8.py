"""
Sử dụng thư viện openpyxl để đọc file excel ở câu trước
"""
from openpyxl import load_workbook


def read_excel(path: str):
	try:
		wb = load_workbook(path)
	except Exception as e:
		print(f"Không thể mở file Excel: {e}")
		return []
	sheet = wb[wb.sheetnames[0]]
	return list(sheet.values)


def print_rows(rows):
	for row in rows:
		for value in row:
			print(value, "\t", end='')
		print()


def main():
	path = 'demo.xlsx'
	rows = read_excel(path)
	if not rows:
		return
	print("Sheets:", rows and 'found')
	print_rows(rows)


if __name__ == "__main__":
	main()